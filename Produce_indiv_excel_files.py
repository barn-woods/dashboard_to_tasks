import openpyxl
from summary_sheet_tests import *
from la_columns import *
from geo_info_from_json import *

"""
Produce xlsx files for each non zero dashboard check relevant to LA tasks, dropping columns missing data in all rows,
columns 'originating_authority', 'further_information_location' and 'geometry'. Large area checks are combined into
one large area task spreadsheet, small area checks are combined into one small area task spreadsheet, and short geo
description checks are combined into one sort geo description task spreadsheet. Despite dropping the geometry column
this script replaces the column with useful info about geometries associated to the charge assigned by the json using
the json_geo_info function in the geo_info_from_json script.
"""


def produce_files(input_file, json_file=None, la_name=None):
    if pd.notnull(json_file):
        json_geo_info(json_file)
        json_df = pd.read_csv('extract_with_modified_columns.csv')
    else:
        json_df = None
    summary_sheet = openpyxl.load_workbook(input_file)['Summary']
    cell_range = list(string.ascii_uppercase[11: 20])  # Define the range of columns containing data
    print(f'extracting info from columns: {cell_range}')

    # Dictionary of empty lists to populate with check names
    dict_check_lists = {
        'small_area_checks': [],
        'large_area_checks': [],
        'short_geo_desc_checks': [],
        'other_checks': [],
    }

    # Dictionary of empty lists to populate with Sheet where the dashboard data is stored
    dict_sheet_lists = {
        'small_area_checks_adv_sheet': [],
        'small_area_checks_comp_sheet': [],
        'large_area_checks_sheet': [],
        'short_geo_desc_checks_sheet': [],
        'other_checks_sheet': []
    }

    for column in cell_range:
        for cell in summary_sheet[column]:
            # exclude null values and exclude values exclude values defined in not_exclude_value function
            if pd.notnull(cell.value) and not_exclude_value(cell.value):
                row_num_of_sheet = int(cell.coordinate[1:])
                if type_of_check(cell.value) == 'small_area_checks':
                    dict_check_lists['small_area_checks'].append(cell.value)
                    column_num_of_adv_sheet = int(string.ascii_uppercase.index(cell.coordinate[0])) + 4
                    dict_sheet_lists['small_area_checks_adv_sheet'].append(summary_sheet.cell(
                        row=row_num_of_sheet, column=column_num_of_adv_sheet).hyperlink.location.split('!')[0])
                    column_num_of_comp_sheet = int(string.ascii_uppercase.index(cell.coordinate[0])) + 6
                    dict_sheet_lists['small_area_checks_comp_sheet'].append(summary_sheet.cell(
                        row=row_num_of_sheet, column=column_num_of_comp_sheet).hyperlink.location.split('!')[0])
                else:
                    dict_check_lists[type_of_check(cell.value)[0]].append(cell.value)
                    column_num_of_sheet = int(string.ascii_uppercase.index(cell.coordinate[0])) + 3
                    dict_sheet_lists[type_of_check(cell.value)[1]].append(summary_sheet.cell(
                        row=row_num_of_sheet, column=column_num_of_sheet).hyperlink.location.split('!')[0])

    # Dictionary of empty lists to populate with data frame names for combined checks
    dict_combined_check_dfs = {
        'small_area_checks_adv_dfs': [],
        'small_area_checks_comp_dfs': [],
        'large_area_checks_dfs': [],
        'short_geo_desc_checks_dfs': []
    }
    # Create a dataframe for the small area checks
    for check, check_sheet in zip(dict_check_lists['small_area_checks'],
                                  dict_sheet_lists['small_area_checks_adv_sheet']):
        check = pd.read_excel(input_file, sheet_name=check_sheet, skiprows=1)
        dict_combined_check_dfs['small_area_checks_adv_dfs'].append(check)
    for check, check_sheet in zip(dict_check_lists['small_area_checks'],
                                  dict_sheet_lists['small_area_checks_comp_sheet']):
        check = pd.read_excel(input_file, sheet_name=check_sheet, skiprows=1)
        dict_combined_check_dfs['small_area_checks_comp_dfs'].append(check)
    small_area_check = pd.concat(dict_combined_check_dfs['small_area_checks_adv_dfs']).append(pd.concat(
        dict_combined_check_dfs['small_area_checks_comp_dfs']))
    if la_name is not None:
        check_name = la_name + '_small_area_check'
    else:
        check_name = 'small_area_check'
    create_excel_sheets(df=small_area_check, df_name=check_name, json_data=json_df, json_present=json_file)

    # Create a dataframe for the large area checks and short geo desc checks
    concat_list = ['large_area_checks', 'short_geo_desc_checks']
    for value in concat_list:
        for check, check_sheet in zip(dict_check_lists[value], dict_sheet_lists[value + '_sheet']):
            check = pd.read_excel(input_file, sheet_name=check_sheet, skiprows=1)
            dict_combined_check_dfs[value + '_dfs'].append(check)
        value_name = value
        value = pd.concat(dict_combined_check_dfs['large_area_checks_dfs'])
        if la_name is not None:
            value.name = la_name + '_' + value_name
        else:
            value.name = value_name
        create_excel_sheets(df=value, df_name=value.name, json_data=json_df, json_present=json_file)

    # Create a dataframes for the other checks
    for check, sheet in zip(dict_check_lists['other_checks'], dict_sheet_lists['other_checks_sheet']):
        check_df = pd.read_excel(input_file, sheet_name=sheet, skiprows=1)
        if la_name is not None:
            check_df.name = la_name + '_' + check
        else:
            check_df.name = check
        create_excel_sheets(df=check_df, df_name=check_df.name, json_data=json_df, json_present=json_file)


"""
Create the excel file form each dataframe.
"""


def create_excel_sheets(df, df_name, json_present, json_data):
    if len(df) > 0:
        print(f'{df_name} length = {len(df)} rows')
        if pd.notnull(json_present):
            df = df.merge(json_data, left_on='originating_authority_charge_identifier',
                          right_on='originating-authority-charge-identifier', how='left').drop(
                columns=['originating-authority-charge-identifier'])
            poly_columns = ['number_of_polygons', 'number_of_points', 'number_of_lines', 'avg_area_(m^3)',
                            'smallest_area_(m^3)', 'largest_area_(m^3)']
            for column in poly_columns:
                if df[column].sum() == 0:
                    df = df.drop(columns=[column])
        if 'distance' in df.columns:
            df['distance'] = df['distance'].astype(float)
            df = df.set_index('distance', drop=False).sort_index(ascending=False)
        df.replace('', None)
        df = df.drop(columns=['originating_authority', 'further_information_location',
                              'geometry']).dropna(axis=1, how='all')
        la_columns_and_format_sheet(df, df_name)
    else:
        print(f'{df_name} length = 0 rows')
