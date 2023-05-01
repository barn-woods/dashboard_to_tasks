import pandas as pd
import string
from openpyxl import load_workbook

"""
Add a column with dropdown options called LA_action, a free text column called LA_comments,
format the column widths to be the same as the max character sting in the column +4 unless this is greater than 35
in which case  a width of 39 characters is used, and freeze the first row.
"""


def la_columns_and_format_sheet(file_instance, file_name):
    file_instance['LA_action'] = None
    file_instance['LA_comment'] = None
    num_cols = len(file_instance.columns)
    num_rows = int(len(file_instance) + 1)
    last_column = string.ascii_uppercase[num_cols - 1]
    idx = file_instance.columns.get_loc('LA_action')
    writer = pd.ExcelWriter(file_name + '.xlsx', engine='xlsxwriter')
    file_instance.to_excel(writer, sheet_name='Sheet1', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    cell_range = string.ascii_uppercase[idx] + '2:' + string.ascii_uppercase[idx] + str(len(file_instance) + 1)
    worksheet.data_validation(
        cell_range, {'validate': 'list', 'source': ['Corrected', 'Cancelled', 'No action required'],
                     'dropdown': True, 'error_title': 'Invalid input',
                     'error_message': 'Please select an action from the dropdown'})
    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(f"A1:{last_column}{num_rows}")
    # worksheet.autofit()
    workbook.close()

    # wb = load_workbook(file_name + '.xlsx')
    # ws = wb['Sheet1']
    # for letter in ['A', last_column]:
    #     for row_number in range(1, ws.max_row + 1):
    #         if len(str(ws[f'{letter}{row_number}'].value)) > 45:
    #             ws.column_dimensions[letter].width = 45
    # wb.close()
